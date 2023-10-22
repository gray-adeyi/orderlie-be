from openpyxl import Workbook
from openpyxl.styles import Font


def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)
    return workbook

def create_sheet(workbook):
    sheet = workbook.active
    return sheet
    


def export_to_excel(connection, query_string, headings, path):
    cursor = connection.cursor()
    cursor.execute(query_string)
    data = cursor.fetchall()
    cursor.close()
    wb = create_workbook(path)
    sheet = create_sheet(wb)
    sheet.row_dimensions[1].font = Font(bold = True)
 
    for colno, heading in enumerate(headings, start = 1):
        sheet.cell(row = 1, column = colno).value = heading


    for rowno, row in enumerate(data, start = 2):
        for colno, cell_value in enumerate(row, start = 1):
            sheet.cell(row = rowno, column = colno).value = cell_value

    wb.save(path)